import re
import logging
import base64
import json
from flask import jsonify, request
from models import db, Location, Resource, location_resource
from datetime import datetime
from sqlalchemy import select, text, and_
import openpyxl
from decimal import Decimal
from sqlalchemy.exc import IntegrityError
from models import ResourceType
from dateutil import parser

logger = logging.getLogger(__name__)

def decode_if_base64(value):
    """
        Decodifica Base64 somente se value for string base64 válida.
        Se for bytes ou None, retorna diretamente.
    """
    if not value:
        return None
    if isinstance(value, bytes):
        return value
    try:
        return base64.b64decode(value)
    except (base64.binascii.Error, ValueError):
        return value.encode() if isinstance(value, str) else value
    
def encode_if_bytes(value):
    if isinstance(value, bytes):
        return base64.b64encode(value).decode('utf-8')
    return value

def decode_file(value):
    if not value:
        return None
    if isinstance(value, bytes):
        return value

    # Tentar Base64
    try:
        return base64.b64decode(value)
    except Exception:
        pass

    # Tentar HEX
    try:
        return bytes.fromhex(value)
    except Exception:
        pass

    # Se não for nenhum formato conhecido
    return None

def extract_base64_from_data_url(data_url):
    """Extrai a parte Base64 de uma data URL (ex: data:image/png;base64,xxxx)"""
    if not data_url:
        return None
    match = re.match(r'^data:[^;]+;base64,(.+)', data_url)
    if match:
        return match.group(1)
    return data_url  # se não for data URL, assume que já é Base64 puro

def extract_mime_and_base64(data_url):
    """Extrai o tipo MIME e a parte Base64 de uma data URL."""
    if not data_url:
        return None, None
    match = re.match(r'^data:([^;]+);base64,(.+)', data_url)
    if match:
        return match.group(1), match.group(2)
    return None, None

def create_data_url(mime_type, base64_data):
    """Cria uma data URL a partir do tipo MIME e dados Base64"""
    return f"data:{mime_type};base64,{base64_data}"

def binary_to_data_url(data, mime_type):
    """Converte bytes em data URL completa."""
    if not data:
        return None
    base64_str = base64.b64encode(data).decode('utf-8')
    return f"data:{mime_type};base64,{base64_str}"

class LocationController:
    @staticmethod
    def get_all():
        try:
            logger.info("[GET ALL] Fetching all locations")

            locations = Location.query.all()
            result = []

            def serialize_resource(link):
                # Imagem principal
                imagem_principal_data = None
                if link.imagem_principal:
                    mime = getattr(link, 'imagem_principal_mime', 'image/jpeg')
                    imagem_principal_data = binary_to_data_url(link.imagem_principal, mime)

                # Imagens adicionais (já é string JSON)
                imagens_data = link.imagens

                # PDF
                anexospdf_data = None
                if link.anexospdf:
                    anexospdf_data = binary_to_data_url(link.anexospdf, 'application/pdf')

                return {
                    'id': link.id,
                    'resource_id': link.resource_id,
                    'name': link.name,
                    'description': link.description,
                    'recebidopor': link.recebidopor,
                    'asset_code': link.asset_code,
                    'budget_to_location': link.budget_to_location,
                    'imagem_principal': imagem_principal_data,
                    'imagens': imagens_data,
                    'anexospdf': anexospdf_data,
                    'datarecepcao': link.datarecepcao.isoformat() if link.datarecepcao else None,
                    'quantity': link.quantity,
                    'status': link.status,
                    'condition': link.condition,
                    # NOVOS CAMPOS
                    'serial_number': link.serial_number,
                    'item_number': link.item_number,
                    'owner': link.owner,
                    'comments': link.comments,
                    'purchase_date': link.purchase_date.isoformat() if link.purchase_date else None,
                    'purchase_cost': str(link.purchase_cost) if link.purchase_cost else None,  # Decimal -> string
                    'inventory_date': link.inventory_date.isoformat() if link.inventory_date else None,
                    'vendor': link.vendor,
                    'project': link.project,
                    'po_number': link.po_number,
                    'observation': link.observation,
                    'createAt': link.createAt.isoformat() if link.createAt else None,
                    'updateAt': link.updateAt.isoformat() if link.updateAt else None,
                }

            for loc in locations:
                stmt = select(location_resource).where(location_resource.c.location_id == loc.id)
                links = db.session.execute(stmt).fetchall()
                resources_data = [serialize_resource(link) for link in links]

                result.append({
                    'id': loc.id,
                    'name': loc.name,
                    'description': loc.description,
                    'latitude': loc.latitude,
                    'longitude': loc.longitude,
                    'responsavel': loc.responsavel,
                    'observacoes': loc.observacoes,
                    'resources': resources_data,
                    'createAt': loc.createAt.isoformat() if loc.createAt else None,
                    'updateAt': loc.updateAt.isoformat() if loc.updateAt else None
                })

            return jsonify(result), 200

        except Exception as e:
            logger.error(f"[GET ALL] Failed to fetch locations: {e}")
            return jsonify({'error': str(e)}), 500

    @staticmethod
    def get_by_id(id):
        try:
            logger.info(f"[GET BY ID] Fetching location ID: {id}")

            loc = Location.query.get(id)
            if not loc:
                return jsonify({'message': 'Location not found'}), 404

            def serialize_resource(link):
                # (mesma função de serialização acima)
                imagem_principal_data = None
                if link.imagem_principal:
                    mime = getattr(link, 'imagem_principal_mime', 'image/jpeg')
                    imagem_principal_data = binary_to_data_url(link.imagem_principal, mime)

                imagens_data = link.imagens
                anexospdf_data = None
                if link.anexospdf:
                    anexospdf_data = binary_to_data_url(link.anexospdf, 'application/pdf')

                return {
                    'id': link.id,
                    'resource_id': link.resource_id,
                    'name': link.name,
                    'description': link.description,
                    'recebidopor': link.recebidopor,
                    'asset_code': link.asset_code,
                    'budget_to_location': link.budget_to_location,
                    'imagem_principal': imagem_principal_data,
                    'imagens': imagens_data,
                    'anexospdf': anexospdf_data,
                    'datarecepcao': link.datarecepcao.isoformat() if link.datarecepcao else None,
                    'quantity': link.quantity,
                    'status': link.status,
                    'condition': link.condition,
                    'serial_number': link.serial_number,
                    'item_number': link.item_number,
                    'owner': link.owner,
                    'comments': link.comments,
                    'purchase_date': link.purchase_date.isoformat() if link.purchase_date else None,
                    'purchase_cost': str(link.purchase_cost) if link.purchase_cost else None,
                    'inventory_date': link.inventory_date.isoformat() if link.inventory_date else None,
                    'vendor': link.vendor,
                    'project': link.project,
                    'po_number': link.po_number,
                    'observation': link.observation,
                    'createAt': link.createAt.isoformat() if link.createAt else None,
                    'updateAt': link.updateAt.isoformat() if link.updateAt else None,
                }

            stmt = select(location_resource).where(location_resource.c.location_id == loc.id)
            links = db.session.execute(stmt).fetchall()
            resources_data = [serialize_resource(link) for link in links]

            return jsonify({
                'id': loc.id,
                'name': loc.name,
                'description': loc.description,
                'latitude': loc.latitude,
                'longitude': loc.longitude,
                'responsavel': loc.responsavel,
                'observacoes': loc.observacoes,
                'resources': resources_data,
                'createAt': loc.createAt.isoformat() if loc.createAt else None,
                'updateAt': loc.updateAt.isoformat() if loc.updateAt else None
            }), 200

        except Exception as e:
            logger.error(f"[GET BY ID] Error fetching location ID {id}: {e}")
            return jsonify({'error': str(e)}), 500

    @staticmethod
    def create():
        try:
            data = request.get_json()
            name = data.get('name')
            description = data.get('description')
            latitude = data.get('latitude')
            longitude = data.get('longitude')
            responsavel = data.get('responsavel', 'DOD')
            observacoes = data.get('observacoes')
            resources = data.get('resources', [])

            if not name or not description:
                return jsonify({'message': 'Missing required data'}), 400

            if responsavel not in Location.get_responsaveis_permitidos():
                return jsonify({'message': f'Responsável deve ser um dos: {", ".join(Location.get_responsaveis_permitidos())}'}), 400

            if Location.query.filter_by(name=name).first():
                return jsonify({'message': 'Location already exists'}), 400

            new_location = Location(
                name=name,
                description=description,
                latitude=float(latitude) if latitude else None,
                longitude=float(longitude) if longitude else None,
                responsavel=responsavel,
                observacoes=observacoes
            )
            db.session.add(new_location)
            db.session.flush()  # Obter ID da nova Location

            # Inserir recursos
            for resource_data in resources:
                # Processar imagem principal
                imagem_principal_data = resource_data.get('imagem_principal')
                imagem_principal_bytes = None
                imagem_principal_mime = None
                if imagem_principal_data:
                    mime, b64 = extract_mime_and_base64(imagem_principal_data)
                    if b64:
                        imagem_principal_bytes = base64.b64decode(b64)
                        imagem_principal_mime = mime

                # Processar imagens adicionais (já é JSON string)
                imagens_json = resource_data.get('imagens')  # ex: '["data:image/png;base64,..."]'

                # Processar PDF
                anexospdf_data = resource_data.get('anexospdf')
                anexospdf_bytes = None
                if anexospdf_data:
                    _, b64 = extract_mime_and_base64(anexospdf_data)
                    if b64:
                        anexospdf_bytes = base64.b64decode(b64)

                # Processar purchase_cost (converter para Decimal)
                purchase_cost = resource_data.get('purchase_cost')
                if purchase_cost is not None:
                    try:
                        purchase_cost = Decimal(str(purchase_cost))
                    except:
                        purchase_cost = None

                db.session.execute(
                    location_resource.insert().values(
                        location_id=new_location.id,
                        resource_id=resource_data.get('resource_id'),
                        quantity=resource_data.get('quantity', 0),
                        status=resource_data.get('status', 'Available'),
                        condition=resource_data.get('condition', 'Good'),
                        name=resource_data.get('name', ''),
                        description=resource_data.get('description'),
                        recebidopor=resource_data.get('recebidopor'),
                        asset_code=resource_data.get('asset_code'),
                        budget_to_location=resource_data.get('budget_to_location'),
                        imagem_principal=imagem_principal_bytes,
                        imagem_principal_mime=imagem_principal_mime,
                        imagens=imagens_json,
                        anexospdf=anexospdf_bytes,
                        datarecepcao=resource_data.get('datarecepcao'),
                        # NOVOS CAMPOS
                        serial_number=resource_data.get('serial_number'),
                        item_number=resource_data.get('item_number'),
                        owner=resource_data.get('owner'),
                        comments=resource_data.get('comments'),
                        purchase_date=resource_data.get('purchase_date'),
                        purchase_cost=purchase_cost,
                        inventory_date=resource_data.get('inventory_date'),
                        vendor=resource_data.get('vendor'),
                        project=resource_data.get('project'),
                        po_number=resource_data.get('po_number'),
                        observation=resource_data.get('observation'),
                    )
                )

            db.session.commit()
            return jsonify({'message': 'Location created successfully'}), 201

        except Exception as e:
            db.session.rollback()
            logger.error(f"[CREATE] Failed to create location: {e}")
            return jsonify({'error': str(e)}), 500

    @staticmethod
    def update(id):
        try:
            loc = Location.query.get(id)
            if not loc:
                return jsonify({'message': 'Location not found'}), 404

            data = request.get_json()
            name = data.get('name')
            description = data.get('description')
            latitude = data.get('latitude')
            longitude = data.get('longitude')
            responsavel = data.get('responsavel')
            observacoes = data.get('observacoes')
            resources = data.get('resources', [])

            if not name or not description:
                return jsonify({'message': 'Missing required data'}), 400

            if responsavel and responsavel not in Location.get_responsaveis_permitidos():
                return jsonify({'message': f'Responsável deve ser um dos: {", ".join(Location.get_responsaveis_permitidos())}'}), 400

            if Location.query.filter(Location.name == name, Location.id != id).first():
                return jsonify({'message': 'Location name already in use'}), 400

            # Atualizar dados da Location
            loc.name = name
            loc.description = description
            loc.latitude = float(latitude) if latitude else None
            loc.longitude = float(longitude) if longitude else None
            if responsavel:
                loc.responsavel = responsavel
            loc.observacoes = observacoes
            loc.updateAt = datetime.utcnow()

            # Remover todos os recursos atuais
            db.session.execute(location_resource.delete().where(location_resource.c.location_id == loc.id))

            # Inserir novamente com os dados atualizados
            for resource_data in resources:
                # Processar imagem principal
                imagem_principal_data = resource_data.get('imagem_principal')
                imagem_principal_bytes = None
                imagem_principal_mime = None
                if imagem_principal_data:
                    mime, b64 = extract_mime_and_base64(imagem_principal_data)
                    if b64:
                        imagem_principal_bytes = base64.b64decode(b64)
                        imagem_principal_mime = mime

                # Processar imagens adicionais
                imagens_json = resource_data.get('imagens')

                # Processar PDF
                anexospdf_data = resource_data.get('anexospdf')
                anexospdf_bytes = None
                if anexospdf_data:
                    _, b64 = extract_mime_and_base64(anexospdf_data)
                    if b64:
                        anexospdf_bytes = base64.b64decode(b64)

                # Processar purchase_cost
                purchase_cost = resource_data.get('purchase_cost')
                if purchase_cost is not None:
                    try:
                        purchase_cost = Decimal(str(purchase_cost))
                    except:
                        purchase_cost = None

                db.session.execute(
                    location_resource.insert().values(
                        location_id=loc.id,
                        resource_id=resource_data.get('resource_id'),
                        quantity=resource_data.get('quantity', 0),
                        status=resource_data.get('status', 'Available'),
                        condition=resource_data.get('condition', 'Good'),
                        name=resource_data.get('name', ''),
                        description=resource_data.get('description'),
                        recebidopor=resource_data.get('recebidopor'),
                        asset_code=resource_data.get('asset_code'),
                        budget_to_location=resource_data.get('budget_to_location'),
                        imagem_principal=imagem_principal_bytes,
                        imagem_principal_mime=imagem_principal_mime,
                        imagens=imagens_json,
                        anexospdf=anexospdf_bytes,
                        datarecepcao=resource_data.get('datarecepcao'),
                        serial_number=resource_data.get('serial_number'),
                        item_number=resource_data.get('item_number'),
                        owner=resource_data.get('owner'),
                        comments=resource_data.get('comments'),
                        purchase_date=resource_data.get('purchase_date'),
                        purchase_cost=purchase_cost,
                        inventory_date=resource_data.get('inventory_date'),
                        vendor=resource_data.get('vendor'),
                        project=resource_data.get('project'),
                        po_number=resource_data.get('po_number'),
                        observation=resource_data.get('observation'),
                    )
                )

            db.session.commit()
            return jsonify({'message': 'Location updated successfully'}), 200

        except Exception as e:
            db.session.rollback()
            logger.error(f"[UPDATE] Failed to update location ID {id}: {e}")
            return jsonify({'error': str(e)}), 500

    @staticmethod
    def delete(id):
        try:
            loc = Location.query.get(id)
            if not loc:
                return jsonify({'message': 'Location not found'}), 404

            db.session.execute(
                location_resource.delete().where(
                    location_resource.c.location_id == id
                )
            )
            db.session.delete(loc)
            db.session.commit()
            return jsonify({'message': 'Location deleted successfully'}), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @staticmethod
    def get_responsaveis():
        try:
            responsaveis = Location.get_responsaveis_permitidos()
            return jsonify({'responsaveis': responsaveis, 'total': len(responsaveis)}), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # 🔹 NOVO MÉTODO: ATUALIZAR APENAS UM RECURSO ESPECÍFICO
    @staticmethod
    def update_resource(location_id, resource_id):
        try:
            data = request.get_json()
            
            # Verificar se o registro existe
            stmt = select(location_resource).where(
                and_(
                    location_resource.c.location_id == location_id,
                    location_resource.c.resource_id == resource_id
                )
            )
            existing = db.session.execute(stmt).fetchone()
            
            if not existing:
                return jsonify({'message': 'Resource not found in this location'}), 404

            # Preparar valores para atualização
            update_values = {
                'quantity': data.get('quantity', existing.quantity),
                'status': data.get('status', existing.status),
                'condition': data.get('condition', existing.condition),
                'name': data.get('name', existing.name),
                'description': data.get('description', existing.description),
                'recebidopor': data.get('recebidopor', existing.recebidopor),
                'asset_code': data.get('asset_code', existing.asset_code),
                'budget_to_location': data.get('budget_to_location', existing.budget_to_location),
                'datarecepcao': data.get('datarecepcao', existing.datarecepcao),
                'updateAt': datetime.utcnow(),
                # NOVOS CAMPOS
                'serial_number': data.get('serial_number', existing.serial_number),
                'item_number': data.get('item_number', existing.item_number),
                'owner': data.get('owner', existing.owner),
                'comments': data.get('comments', existing.comments),
                'purchase_date': data.get('purchase_date', existing.purchase_date),
                'inventory_date': data.get('inventory_date', existing.inventory_date),
                'vendor': data.get('vendor', existing.vendor),
                'project': data.get('project', existing.project),
                'po_number': data.get('po_number', existing.po_number),
                'observation': data.get('observation', existing.observation),
            }

            # Processar purchase_cost separadamente (Decimal)
            purchase_cost = data.get('purchase_cost')
            if purchase_cost is not None:
                try:
                    update_values['purchase_cost'] = Decimal(str(purchase_cost))
                except:
                    pass
            else:
                update_values['purchase_cost'] = existing.purchase_cost

            # Processar imagem principal se fornecida
            imagem_principal_data = data.get('imagem_principal')
            if imagem_principal_data:
                mime, b64 = extract_mime_and_base64(imagem_principal_data)
                if b64:
                    update_values['imagem_principal'] = base64.b64decode(b64)
                    update_values['imagem_principal_mime'] = mime
            elif 'imagem_principal' in data and data['imagem_principal'] is None:
                # Se o campo for enviado como null, remover a imagem
                update_values['imagem_principal'] = None
                update_values['imagem_principal_mime'] = None

            # Processar imagens adicionais
            if 'imagens' in data:
                update_values['imagens'] = data['imagens']

            # Processar PDF
            anexospdf_data = data.get('anexospdf')
            if anexospdf_data:
                _, b64 = extract_mime_and_base64(anexospdf_data)
                if b64:
                    update_values['anexospdf'] = base64.b64decode(b64)
            elif 'anexospdf' in data and data['anexospdf'] is None:
                update_values['anexospdf'] = None

            # Atualizar
            update_stmt = location_resource.update().where(
                and_(
                    location_resource.c.location_id == location_id,
                    location_resource.c.resource_id == resource_id
                )
            ).values(**update_values)

            db.session.execute(update_stmt)
            db.session.commit()

            return jsonify({'message': 'Resource updated successfully'}), 200
        except Exception as e:
            db.session.rollback()
            logger.error(f"[UPDATE RESOURCE] Failed to update resource {resource_id} in location {location_id}: {e}")
            return jsonify({'error': str(e)}), 500

    # 🔹 NOVO MÉTODO: REMOVER APENAS UM RECURSO ESPECÍFICO
    @staticmethod
    def delete_resource(location_id, resource_id):
        try:
            delete_stmt = location_resource.delete().where(
                and_(
                    location_resource.c.location_id == location_id,
                    location_resource.c.resource_id == resource_id
                )
            )
            result = db.session.execute(delete_stmt)
            db.session.commit()

            if result.rowcount == 0:
                return jsonify({'message': 'Resource not found in this location'}), 404

            return jsonify({'message': 'Resource removed from location successfully'}), 200
        except Exception as e:
            db.session.rollback()
            logger.error(f"[DELETE RESOURCE] Failed to delete resource {resource_id} from location {location_id}: {e}")
            return jsonify({'error': str(e)}), 500
        
    @staticmethod
    def upload_excel():
        """Endpoint para upload de arquivo Excel com dados de ativos"""
        import openpyxl
        from decimal import Decimal
        from datetime import datetime
        from sqlalchemy.exc import IntegrityError
        from models import ResourceType
        from dateutil import parser

        if 'excel_file' not in request.files:
            return jsonify({'error': 'No file part'}), 400

        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'error': 'No selected file'}), 400

        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'File must be .xlsx'}), 400

        try:
            workbook = openpyxl.load_workbook(file, data_only=True)
            sheet = workbook.active
            logger.info(f"[UPLOAD] Planilha carregada, total de linhas: {sheet.max_row}, colunas: {sheet.max_column}")
        except Exception as e:
            logger.error(f"[UPLOAD] Erro ao ler arquivo Excel: {e}")
            return jsonify({'error': f'Could not read Excel file: {str(e)}'}), 400

        # Cabeçalhos
        headers = [cell.value for cell in sheet[1]]
        logger.info(f"[UPLOAD] Cabeçalhos encontrados: {headers}")

        col_index = {
            'asset_number': None,
            'serial_number': None,
            'status': None,
            'item_number': None,
            'description': None,
            'site': None,
            'location': None,
            'owner': None,
            'comments': None,
            'purchase_date': None,
            'purchase_cost': None,
            'condition': None,
            'inventory_date': None,
            'vendor': None,
            'donor': None,
            'project': None,
            'po_number': None,
            'observation': None,
        }

        # Mapear colunas
        for idx, h in enumerate(headers):
            if h is None:
                continue
            h_str = str(h).strip()
            h_lower = h_str.lower()
            if 'asset number' in h_lower:
                col_index['asset_number'] = idx
                logger.info(f"[UPLOAD] Asset Number -> coluna {idx}")
            elif 'serial number' in h_lower:
                col_index['serial_number'] = idx
                logger.info(f"[UPLOAD] Serial Number -> coluna {idx}")
            elif 'status' in h_lower:
                col_index['status'] = idx
                logger.info(f"[UPLOAD] Status -> coluna {idx}")
            elif 'item number' in h_lower:
                col_index['item_number'] = idx
                logger.info(f"[UPLOAD] Item Number -> coluna {idx}")
            elif 'description' in h_lower:
                col_index['description'] = idx
                logger.info(f"[UPLOAD] Description -> coluna {idx}")
            elif 'site' in h_lower:
                col_index['site'] = idx
                logger.info(f"[UPLOAD] Site -> coluna {idx}")
            elif 'location' in h_lower or 'local' in h_lower:
                col_index['location'] = idx
                logger.info(f"[UPLOAD] Location -> coluna {idx}")
            elif 'owner' in h_lower:
                col_index['owner'] = idx
                logger.info(f"[UPLOAD] Owner -> coluna {idx}")
            elif 'comments' in h_lower:
                col_index['comments'] = idx
                logger.info(f"[UPLOAD] Comments -> coluna {idx}")
            elif 'purchase date' in h_lower:
                col_index['purchase_date'] = idx
                logger.info(f"[UPLOAD] Purchase Date -> coluna {idx}")
            elif 'purchase cost' in h_lower:
                col_index['purchase_cost'] = idx
                logger.info(f"[UPLOAD] Purchase Cost -> coluna {idx}")
            elif 'condition' in h_lower:
                col_index['condition'] = idx
                logger.info(f"[UPLOAD] Condition -> coluna {idx}")
            elif 'inventory date' in h_lower:
                col_index['inventory_date'] = idx
                logger.info(f"[UPLOAD] Inventory Date -> coluna {idx}")
            elif 'vendor' in h_lower:
                col_index['vendor'] = idx
                logger.info(f"[UPLOAD] Vendor -> coluna {idx}")
            elif 'donor' in h_lower:
                col_index['donor'] = idx
                logger.info(f"[UPLOAD] Donor -> coluna {idx}")
            elif 'project' in h_lower or 'award name' in h_lower:
                col_index['project'] = idx
                logger.info(f"[UPLOAD] Project -> coluna {idx}")
            elif 'po#' in h_lower or 'po' in h_lower:
                col_index['po_number'] = idx
                logger.info(f"[UPLOAD] PO# -> coluna {idx}")
            elif 'observation' in h_lower:
                col_index['observation'] = idx
                logger.info(f"[UPLOAD] Observation -> coluna {idx}")

        required = ['description', 'location']
        missing = [req for req in required if col_index[req] is None]
        if missing:
            error_msg = f"Colunas obrigatórias não encontradas: {', '.join(missing)}. Cabeçalhos: {headers}"
            logger.error(f"[UPLOAD] {error_msg}")
            return jsonify({'error': error_msg}), 400

        # --- Garantir que existe um ResourceType com ID = 2 (recurso material) ---
        material_resource_type = ResourceType.query.get(2)
        if not material_resource_type:
            material_resource_type = ResourceType(id=2, name='Material', description='Recursos materiais')
            db.session.add(material_resource_type)
            db.session.commit()
            logger.info("[UPLOAD] ResourceType com ID 2 criado: Material")
        else:
            logger.info("[UPLOAD] ResourceType com ID 2 já existe: Material")

        created_count = 0
        skipped_count = 0
        errors = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            logger.info(f"[UPLOAD] Processando linha {row_idx}: {row}")

            if all(cell is None or str(cell).strip() == '' for cell in row):
                logger.info(f"[UPLOAD] Linha {row_idx} vazia, ignorando")
                continue

            # --- Location (ID ou nome) ---
            loc_val = row[col_index['location']]
            logger.info(f"[UPLOAD] Linha {row_idx} - Valor bruto da localização: '{loc_val}' (tipo: {type(loc_val)})")

            location = None
            try:
                loc_id = int(loc_val) if loc_val is not None else None
                if loc_id is not None:
                    location = Location.query.get(loc_id)
                    if location:
                        logger.info(f"[UPLOAD] Linha {row_idx} - Localização encontrada por ID {loc_id}: {location.name}")
            except (ValueError, TypeError):
                pass

            if location is None and loc_val:
                loc_name = str(loc_val).strip()
                location = Location.query.filter(Location.name.ilike(loc_name)).first()
                if location:
                    logger.info(f"[UPLOAD] Linha {row_idx} - Localização encontrada por nome '{loc_name}': {location.name} (ID {location.id})")

            if not location:
                errors.append(f"Linha {row_idx}: localização não encontrada (valor: {loc_val})")
                skipped_count += 1
                continue

            # --- Resource (tipo) ---
            desc_val = row[col_index['description']]
            if not desc_val:
                errors.append(f"Linha {row_idx}: sem descrição do recurso")
                skipped_count += 1
                continue
            desc = str(desc_val).strip()

            resource = Resource.query.filter_by(name=desc).first()
            if not resource:
                resource = Resource(
                    name=desc,
                    description=desc,
                    resourcetypeId=2
                )
                db.session.add(resource)
                db.session.flush()
                logger.info(f"[UPLOAD] Linha {row_idx} - Novo recurso criado: {desc} (ID {resource.id}, resourcetypeId=2)")
            else:
                logger.info(f"[UPLOAD] Linha {row_idx} - Recurso existente: {desc} (ID {resource.id})")

            # --- Tratar owner (se for apenas espaço, transforma em None) ---
            owner_val = row[col_index['owner']] if col_index['owner'] is not None else None
            owner = None
            if owner_val and str(owner_val).strip():
                owner = str(owner_val).strip()

            # --- Tratar status ---
            status_val = row[col_index['status']] if col_index['status'] is not None else None
            status = 'Available'
            if status_val and str(status_val).strip():
                status = str(status_val).strip()

            # --- Tratar condition ---
            condition_val = row[col_index['condition']] if col_index['condition'] is not None else None
            condition_map = {
                'Good': 'Good', 'good': 'Good',
                'Fair': 'Fair', 'fair': 'Fair',
                'Poor': 'Poor', 'poor': 'Poor',
                'Damaged': 'Damaged', 'damaged': 'Damaged',
                'Needs Repair': 'Needs Repair', 'needs repair': 'Needs Repair',
            }
            condition = condition_map.get(condition_val, 'Good') if condition_val else 'Good'

            # --- Asset code como string ---
            asset_code_raw = row[col_index['asset_number']] if col_index['asset_number'] is not None else None
            asset_code_str = None
            if asset_code_raw is not None:
                asset_code_str = str(asset_code_raw).strip()
                if asset_code_str == '':
                    asset_code_str = None

            # --- Donor -> adicionar aos comments ---
            comments_val = row[col_index['comments']] if col_index['comments'] is not None else None
            donor = row[col_index['donor']] if col_index['donor'] is not None else None
            comments = comments_val if comments_val else ''
            if donor:
                donor_str = f"Donor: {donor}"
                if comments:
                    comments = f"{comments} | {donor_str}"
                else:
                    comments = donor_str
            comments = comments if comments else None

            # --- Dados do location_resource ---
            data = {
                'location_id': location.id,
                'resource_id': resource.id,
                'name': resource.name,                      # <--- obrigatório
                'quantity': 1,                              # padrão
                'status': status,
                'condition': condition,
                'asset_code': asset_code_str,
                'serial_number': row[col_index['serial_number']] if col_index['serial_number'] is not None else None,
                'item_number': row[col_index['item_number']] if col_index['item_number'] is not None else None,
                'owner': owner,
                'comments': comments,
                'purchase_date': row[col_index['purchase_date']] if col_index['purchase_date'] is not None else None,
                'purchase_cost': row[col_index['purchase_cost']] if col_index['purchase_cost'] is not None else None,
                'inventory_date': row[col_index['inventory_date']] if col_index['inventory_date'] is not None else None,
                'vendor': row[col_index['vendor']] if col_index['vendor'] is not None else None,
                'project': row[col_index['project']] if col_index['project'] is not None else None,
                'po_number': row[col_index['po_number']] if col_index['po_number'] is not None else None,
                'observation': row[col_index['observation']] if col_index['observation'] is not None else None,
            }

            # --- Converter datas com dateutil.parser ---
            for date_field in ['purchase_date', 'inventory_date']:
                val = data.get(date_field)
                if val and not isinstance(val, datetime):
                    try:
                        data[date_field] = parser.parse(str(val))
                    except Exception as e:
                        logger.warning(f"[UPLOAD] Linha {row_idx} - Erro ao converter data {date_field}: {val} - {e}")
                        data[date_field] = None

            # --- Converter purchase_cost para Decimal ---
            if data['purchase_cost']:
                try:
                    data['purchase_cost'] = Decimal(str(data['purchase_cost']))
                except Exception as e:
                    logger.warning(f"[UPLOAD] Linha {row_idx} - Erro ao converter purchase_cost: {data['purchase_cost']} - {e}")
                    data['purchase_cost'] = None

            # --- Evitar duplicatas por asset_code ---
            if asset_code_str:
                existing = db.session.query(location_resource).filter(
                    location_resource.c.asset_code == asset_code_str
                ).first()
                if existing:
                    errors.append(f"Linha {row_idx}: ativo com código {asset_code_str} já existe. Ignorada.")
                    skipped_count += 1
                    continue

            # --- Inserir no location_resource ---
            try:
                stmt = location_resource.insert().values(**data)
                db.session.execute(stmt)
                db.session.commit()
                created_count += 1
                logger.info(f"[UPLOAD] Linha {row_idx} - Recurso inserido com sucesso.")
            except IntegrityError as e:
                db.session.rollback()
                errors.append(f"Linha {row_idx}: erro de integridade - {str(e)}")
                logger.error(f"[UPLOAD] Erro de integridade linha {row_idx}: {e}")
                skipped_count += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f"Linha {row_idx}: erro ao criar recurso - {str(e)}")
                logger.error(f"[UPLOAD] Erro ao inserir linha {row_idx}: {e}")
                skipped_count += 1

        logger.info(f"[UPLOAD] Resumo: criados={created_count}, ignorados={skipped_count}, erros={len(errors)}")
        return jsonify({
            'created': created_count,
            'skipped': skipped_count,
            'errors': errors[:20]
        })